"""
Input profiling and output export.

The profiler is deliberately thorough because the brief's first requirement is
'input analysis': it reports column inference, missing values, duplicates, encoding
damage and placeholder density *before* anything is enriched, so a bad feed is
visible immediately rather than after a 1,000-row run.
"""
from __future__ import annotations

import collections
import csv
import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from backend.normalization.text import is_placeholder, normalize_whitespace, repair_mojibake
from backend.validation.schema import OutputSchema

try:
    from openpyxl import Workbook, load_workbook
    _HAVE_XLSX = True
except Exception:                        # pragma: no cover
    _HAVE_XLSX = False


# ---------------------------------------------------------------------------
# Reading
# ---------------------------------------------------------------------------
def read_table(path: Path, sheet: Optional[str] = None) -> List[Dict[str, str]]:
    """Read a CSV or XLSX into dicts, tolerating BOMs and blank trailing rows."""
    p = Path(path)
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        return _read_xlsx(p, sheet)
    return _read_csv(p)


def _read_csv(path: Path) -> List[Dict[str, str]]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:                                                   # pragma: no cover
        text = raw.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(text, newline=""))
    out: List[Dict[str, str]] = []
    for row in reader:
        clean = {(k or "").strip(): ("" if v is None else str(v))
                 for k, v in row.items() if k is not None}
        if any(v.strip() for v in clean.values()):
            out.append(clean)
    return out


def _read_xlsx(path: Path, sheet: Optional[str]) -> List[Dict[str, str]]:
    if not _HAVE_XLSX:
        raise RuntimeError("openpyxl is required to read XLSX input")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    try:
        header = ["" if c is None else str(c).strip() for c in next(rows)]
    except StopIteration:
        wb.close()
        return []
    out: List[Dict[str, str]] = []
    for r in rows:
        values = ["" if c is None else str(c) for c in r]
        rec = {header[i]: (values[i] if i < len(values) else "")
               for i in range(len(header)) if header[i]}
        if any(v.strip() for v in rec.values()):
            out.append(rec)
    wb.close()
    return out


def sheet_names(path: Path) -> List[str]:
    p = Path(path)
    if p.suffix.lower() not in (".xlsx", ".xlsm") or not _HAVE_XLSX:
        return []
    wb = load_workbook(p, read_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


# ---------------------------------------------------------------------------
# Profiling
# ---------------------------------------------------------------------------
_MOJIBAKE_RE = re.compile(r"Ã.|Â.|â€|â„")


@dataclass
class ColumnProfile:
    name: str
    non_empty: int = 0
    unique: int = 0
    placeholder: int = 0
    max_length: int = 0
    inferred_type: str = "text"
    top_values: List[Tuple[str, int]] = field(default_factory=list)

    @property
    def fill_rate(self) -> float:
        return 0.0

    def as_dict(self, total: int) -> Dict[str, Any]:
        effective = self.non_empty - self.placeholder
        return {
            "name": self.name,
            "non_empty": self.non_empty,
            "placeholder": self.placeholder,
            "effective_values": effective,
            "fill_rate": round(effective / total, 4) if total else 0.0,
            "unique": self.unique,
            "max_length": self.max_length,
            "inferred_type": self.inferred_type,
            "top_values": [{"value": v, "count": c} for v, c in self.top_values],
        }


@dataclass
class InputProfile:
    path: str
    row_count: int
    columns: List[ColumnProfile]
    duplicate_rows: int
    duplicate_part_numbers: List[Tuple[str, int]]
    encoding_issues: int
    warnings: List[str] = field(default_factory=list)

    def as_dict(self) -> Dict[str, Any]:
        return {
            "path": self.path,
            "row_count": self.row_count,
            "column_count": len(self.columns),
            "columns": [c.as_dict(self.row_count) for c in self.columns],
            "duplicate_rows": self.duplicate_rows,
            "duplicate_part_numbers": [{"value": v, "count": c}
                                       for v, c in self.duplicate_part_numbers],
            "encoding_issues": self.encoding_issues,
            "warnings": list(self.warnings),
        }


def _infer_type(values: Sequence[str]) -> str:
    if not values:
        return "empty"
    sample = values[:400]
    if all(re.fullmatch(r"[+-]?\d+", v) for v in sample):
        return "integer"
    if all(re.fullmatch(r"[+-]?\d*\.?\d+", v) for v in sample):
        return "decimal"
    if all(re.match(r"https?://", v) for v in sample):
        return "url"
    if len({v.lower() for v in sample}) <= max(2, len(sample) // 20):
        return "categorical"
    return "text"


def profile_input(rows: Sequence[Dict[str, Any]], path: str = "",
                  key_column: str = "Mfg_Part_Num") -> InputProfile:
    total = len(rows)
    columns = list(rows[0].keys()) if rows else []
    profiles: List[ColumnProfile] = []
    encoding_issues = 0

    for col in columns:
        values: List[str] = []
        placeholder = 0
        for r in rows:
            v = r.get(col)
            s = "" if v is None else str(v).strip()
            if not s:
                continue
            if _MOJIBAKE_RE.search(s):
                encoding_issues += 1
            if is_placeholder(s):
                placeholder += 1
                continue
            values.append(s)
        cp = ColumnProfile(
            name=col,
            non_empty=len(values) + placeholder,
            unique=len(set(values)),
            placeholder=placeholder,
            max_length=max((len(v) for v in values), default=0),
            inferred_type=_infer_type(values),
            top_values=collections.Counter(values).most_common(5),
        )
        profiles.append(cp)

    seen: collections.Counter = collections.Counter()
    for r in rows:
        seen[tuple(sorted((k, str(v)) for k, v in r.items()))] += 1
    duplicate_rows = sum(c - 1 for c in seen.values() if c > 1)

    key_counts = collections.Counter(
        str(r.get(key_column, "")).strip() for r in rows if str(r.get(key_column, "")).strip())
    dup_keys = [(k, c) for k, c in key_counts.most_common(20) if c > 1]

    warnings: List[str] = []
    if not rows:
        warnings.append("input is empty")
    if key_column not in columns:
        warnings.append(f"expected key column '{key_column}' is missing")
    if duplicate_rows:
        warnings.append(f"{duplicate_rows} exactly duplicated row(s)")
    if dup_keys:
        warnings.append(f"{len(dup_keys)} part number(s) appear more than once")
    if encoding_issues:
        warnings.append(f"{encoding_issues} cell(s) show mojibake; encoding repair applied")
    for cp in profiles:
        if total and cp.placeholder / total > 0.5:
            warnings.append(
                f"column '{cp.name}' is {cp.placeholder / total:.0%} placeholder sentinels")

    return InputProfile(path or "", total, profiles, duplicate_rows, dup_keys,
                        encoding_issues, warnings)


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------
def write_csv(rows: Sequence[Dict[str, str]], schema: OutputSchema, path: Path) -> Path:
    """Write the delivery-format CSV with the exact template header and order."""
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(schema.headers)
        for row in rows:
            writer.writerow(schema.order(row))
    return p


def write_xlsx(rows: Sequence[Dict[str, str]], schema: OutputSchema, path: Path,
               sheet_name: str = "Delivery Format") -> Path:
    if not _HAVE_XLSX:
        raise RuntimeError("openpyxl is required to write XLSX output")
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(sheet_name[:31])
    ws.append(schema.headers)
    for row in rows:
        ws.append(schema.order(row))
    wb.save(str(p))
    return p


def write_review_csv(products: Sequence[Any], path: Path) -> Path:
    """Flat, actionable review queue: one row per flagged issue."""
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["Row", "Part Number", "Description", "Status", "Confidence",
                    "Issue", "Field", "Detail", "Suggested Value"])
        for prod in products:
            for flag in prod.review_flags:
                w.writerow([
                    prod.row_index,
                    prod.mpn.value or "",
                    (prod.cleaned.get("Part_Desc") or "")[:120],
                    prod.status.value,
                    f"{prod.confidence:.2f}",
                    flag.reason, flag.field, flag.detail,
                    flag.suggested_value or "",
                ])
    return p
