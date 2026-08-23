"""
Output schema handling.

The 252 delivery-format headers are **never hard-coded**.  They are read from the
supplied delivery-format template at start-up, so if Unilog ships a revised template
with extra columns the pipeline adopts it without a code change.

``OutputSchema`` also parses the column *families* -- ``ATTRIBUTE_LABEL n`` /
``ATTRIBUTE_VALUE n`` / ``ATTRIBUTE_UOM n``, ``ITEM_FEATURES_n``, ``Ref URL n``,
``Alternate Image n`` -- because their arity (how many slots exist) is a property of
the template, not of the code.
"""
from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

try:
    from openpyxl import load_workbook
    _HAVE_XLSX = True
except Exception:                        # pragma: no cover
    _HAVE_XLSX = False

# Column-family patterns: name -> (regex with an index group)
FAMILY_PATTERNS: Dict[str, str] = {
    "attribute_label": r"^ATTRIBUTE_LABEL\s*(\d+)$",
    "attribute_value": r"^ATTRIBUTE_VALUE\s*(\d+)$",
    "attribute_uom": r"^ATTRIBUTE_UOM\s*(\d+)$",
    "item_features": r"^ITEM_FEATURES_(\d+)$",
    "ref_url": r"^Ref URL\s*(\d+)$",
    "alternate_image": r"^Alternate Image\s*(\d+)$",
    "video_link": r"^Video Link\s*(\d+)$",
    "sds": r"^SDS_(\d+)$",
}


@dataclass
class OutputSchema:
    """The delivery-format column contract."""
    headers: List[str] = field(default_factory=list)
    families: Dict[str, List[Tuple[int, str]]] = field(default_factory=dict)
    source: str = ""

    # -- construction -----------------------------------------------------
    @classmethod
    def from_file(cls, path: Path) -> "OutputSchema":
        """Read headers from a delivery-format CSV or XLSX template."""
        p = Path(path)
        if p.suffix.lower() in (".xlsx", ".xlsm"):
            headers = cls._headers_from_xlsx(p)
        else:
            headers = cls._headers_from_csv(p)
        return cls.from_headers(headers, source=p.name)

    @staticmethod
    def _headers_from_csv(path: Path) -> List[str]:
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            for row in csv.reader(fh):
                return [c.strip() for c in row]
        return []

    @staticmethod
    def _headers_from_xlsx(path: Path) -> List[str]:
        if not _HAVE_XLSX:
            raise RuntimeError("openpyxl is required to read an XLSX template")
        wb = load_workbook(path, read_only=True, data_only=True)
        # Prefer a sheet that looks like the delivery format.
        name = next((n for n in wb.sheetnames if "deliver" in n.lower()), wb.sheetnames[0])
        ws = wb[name]
        headers: List[str] = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(c).strip() if c is not None else "" for c in row]
            break
        wb.close()
        return headers

    @classmethod
    def from_headers(cls, headers: Sequence[str], source: str = "") -> "OutputSchema":
        hdrs = [h for h in (str(h or "").strip() for h in headers)]
        families: Dict[str, List[Tuple[int, str]]] = {k: [] for k in FAMILY_PATTERNS}
        for h in hdrs:
            for fam, pat in FAMILY_PATTERNS.items():
                m = re.match(pat, h, re.IGNORECASE)
                if m:
                    families[fam].append((int(m.group(1)), h))
                    break
        for fam in families:
            families[fam].sort(key=lambda t: t[0])
        return cls(headers=hdrs, families=families, source=source)

    # -- queries ----------------------------------------------------------
    def __len__(self) -> int:
        return len(self.headers)

    def has(self, name: str) -> bool:
        return name in self.headers

    def family_size(self, family: str) -> int:
        return len(self.families.get(family, ()))

    def family_columns(self, family: str) -> List[str]:
        return [name for _idx, name in self.families.get(family, ())]

    @property
    def attribute_slots(self) -> int:
        """Number of ATTRIBUTE_* triples the template exposes."""
        return min(self.family_size("attribute_label"),
                   self.family_size("attribute_value"),
                   self.family_size("attribute_uom"))

    def blank_row(self) -> Dict[str, str]:
        return {h: "" for h in self.headers}

    def order(self, record: Dict[str, str]) -> List[str]:
        """Project a dict onto the exact template column order."""
        return [str(record.get(h, "") or "") for h in self.headers]

    def unknown_columns(self, record: Dict[str, str]) -> List[str]:
        known = set(self.headers)
        return [k for k in record if k not in known]

    def summary(self) -> Dict[str, object]:
        return {
            "source": self.source,
            "column_count": len(self.headers),
            "attribute_slots": self.attribute_slots,
            "item_feature_slots": self.family_size("item_features"),
            "ref_url_slots": self.family_size("ref_url"),
            "alternate_image_slots": self.family_size("alternate_image"),
        }


def load_default_schema(candidates: Sequence[Path]) -> Optional[OutputSchema]:
    """First readable template wins; returns None when none is available."""
    for c in candidates:
        p = Path(c)
        if p.exists():
            try:
                schema = OutputSchema.from_file(p)
                if schema.headers:
                    return schema
            except Exception:
                continue
    return None
